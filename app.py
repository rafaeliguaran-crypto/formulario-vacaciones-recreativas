from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

ARCHIVO_EXCEL = "inscripciones_vacaciones.xlsx"

HORARIOS = [
    "Martes 16 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Martes 16 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Miércoles 17 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Miércoles 17 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Jueves 18 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Jueves 18 de Junio entre las 2:00 p. m. y 5:00 p. m.",
    "Viernes 19 de Junio entre las 9:00 a. m. y 12:00 p. m.",
    "Viernes 19 de Junio entre las 2:00 p. m. y 5:00 p. m."
]

ENCABEZADOS = [
    "ID",
    "Autorización Datos",
    "Autorización Imágenes",
    "Tipo Documento",
    "Número Documento",
    "Nombre Completo",
    "Celular",
    "Correo",
    "Comunidad",
    "Espacio Adecuado",
    "Horario"
]

def limpiar(texto):
    if texto is None:
        return ""
    return str(texto).strip()

def crear_excel_si_no_existe():
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscripciones"
        ws.append(ENCABEZADOS)
        wb.save(ARCHIVO_EXCEL)
        wb.close()

def leer_registros():
    crear_excel_si_no_existe()

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Inscripciones"]

    registros = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if any(fila):
            registros.append(fila)

    wb.close()
    return registros

def obtener_horarios_ocupados():
    registros = leer_registros()
    horarios_ocupados = []

    for fila in registros:
        horario = limpiar(fila[10])
        if horario:
            horarios_ocupados.append(horario)

    return horarios_ocupados

def obtener_horarios_disponibles():
    horarios_ocupados = obtener_horarios_ocupados()

    return [
        horario for horario in HORARIOS
        if horario not in horarios_ocupados
    ]

def guardar_registro(datos):
    crear_excel_si_no_existe()

    horario_seleccionado = limpiar(datos[-1])

    if horario_seleccionado in obtener_horarios_ocupados():
        return False

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Inscripciones"]

    nuevo_id = ws.max_row
    datos_limpios = tuple(limpiar(dato) for dato in datos)

    ws.append((nuevo_id,) + datos_limpios)

    wb.save(ARCHIVO_EXCEL)
    wb.close()

    return True

@app.route("/", methods=["GET", "POST"])
def formulario():
    if request.method == "POST":
        datos = (
            request.form.get("autorizacion_datos"),
            request.form.get("autorizacion_imagenes"),
            request.form.get("tipo_documento"),
            request.form.get("numero_documento"),
            request.form.get("nombre_completo"),
            request.form.get("celular"),
            request.form.get("correo"),
            request.form.get("comunidad"),
            request.form.get("espacio_adecuado"),
            request.form.get("horario")
        )

        if guardar_registro(datos):
            return redirect(url_for("gracias"))

        return "Ese horario ya fue seleccionado. Por favor vuelve al formulario y escoge otro horario disponible."

    horarios_disponibles = obtener_horarios_disponibles()
    return render_template("index.html", horarios=horarios_disponibles)

@app.route("/gracias")
def gracias():
    return render_template("gracias.html")

@app.route("/registros")
def registros():
    datos = leer_registros()
    return render_template("registros.html", datos=datos)

@app.route("/debug")
def debug():
    return {
        "ocupados": obtener_horarios_ocupados(),
        "disponibles": obtener_horarios_disponibles()
    }

if __name__ == "__main__":
    app.run(debug=True)